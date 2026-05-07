from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime
from io import BytesIO
from typing import Any

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.security import require_active_user
from app.models.church import (
    ActivityType,
    Attendance,
    IcareGroup,
    IcareMember,
    Member,
)

router = APIRouter(prefix="/api/reports", tags=["Reports"])


# ── Location → activity types & home churches ─────────────────────────────────

LOCATION_OPTIONS = [
    {"value": "taipei",   "label": "Taipei"},
    {"value": "zhongli",  "label": "Zhongli"},
    {"value": "combined", "label": "Combined"},
]

LOCATION_ACTIVITIES = {
    "taipei":   ["Ibadah Taipei"],
    "zhongli":  ["Ibadah Zhongli"],
    "combined": ["Ibadah Taipei", "Ibadah Zhongli"],
}

LOCATION_HOME_CHURCHES = {
    "taipei":   ["IFGF Taipei"],
    "zhongli":  ["IFGF Zhongli"],
    "combined": ["IFGF Taipei", "IFGF Zhongli"],
}

LOCATION_SHORT = {"taipei": "TPE", "zhongli": "ZL", "combined": "Combined"}


CATEGORY_LABEL = {
    "professional": "Adult",
    "family":       "Adult",
    "student":      "College",
    "teenager":     "Teens and Youth",
    "children":     "Kids",
}
CATEGORY_ORDER = ["Adult", "College", "Teens and Youth", "Kids"]


# ── Report type metadata ──────────────────────────────────────────────────────

def _current_year() -> int:
    return date.today().year


REPORT_TYPES = [
    {
        "type": "sunday_attendance_roster",
        "label": "Sunday Service – Yearly Roster",
        "description": "Per-member attendance grid for each Sunday in the year.",
        "params": [
            {
                "key": "location",
                "label": "Location",
                "type": "select",
                "options": LOCATION_OPTIONS,
                "default": "taipei",
            },
            {
                "key": "year",
                "label": "Year",
                "type": "number",
                "default": None,  # filled on the fly
                "min": 2000,
                "max": 2100,
            },
        ],
    },
    {
        "type": "sunday_attendance_summary",
        "label": "Sunday Service – Monthly & Quarterly Summary",
        "description": "Average weekly headcount per category, by month and quarter.",
        "params": [
            {
                "key": "location",
                "label": "Location",
                "type": "select",
                "options": LOCATION_OPTIONS,
                "default": "taipei",
            },
            {
                "key": "year",
                "label": "Year",
                "type": "number",
                "default": None,
                "min": 2000,
                "max": 2100,
            },
        ],
    },
]


@router.get("/types")
def list_report_types(_=Depends(require_active_user)):
    """Metadata for available report types and their parameter schemas."""
    year = _current_year()
    out = []
    for t in REPORT_TYPES:
        params = []
        for p in t["params"]:
            params.append({**p, "default": year if p["key"] == "year" else p.get("default")})
        out.append({**t, "params": params})
    return out


# ── Export request schema ─────────────────────────────────────────────────────

class ReportItem(BaseModel):
    type: str
    params: dict[str, Any] = Field(default_factory=dict)


class ReportExportRequest(BaseModel):
    reports: list[ReportItem]


# ── Helpers ───────────────────────────────────────────────────────────────────

def _validate_location(loc: str) -> str:
    if loc not in LOCATION_ACTIVITIES:
        raise HTTPException(status_code=400, detail=f"Invalid location: {loc}")
    return loc


def _validate_year(year: Any) -> int:
    try:
        y = int(year)
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail=f"Invalid year: {year}")
    if y < 2000 or y > 2100:
        raise HTTPException(status_code=400, detail=f"Year out of range: {y}")
    return y


def _activity_ids_for_location(db: Session, location: str) -> list[int]:
    names = LOCATION_ACTIVITIES[location]
    rows = db.query(ActivityType.id).filter(ActivityType.name.in_(names)).all()
    return [r[0] for r in rows]


def _kategori(member_category: str | None) -> str:
    if not member_category:
        return "Unknown"
    return CATEGORY_LABEL.get(member_category, member_category)


def _unique_sheet_name(wb: Workbook, base: str) -> str:
    base = base[:31]
    if base not in wb.sheetnames:
        return base
    for i in range(2, 100):
        suffix = f" ({i})"
        candidate = base[: 31 - len(suffix)] + suffix
        if candidate not in wb.sheetnames:
            return candidate
    return base + "_x"


# ── Sheet builders ────────────────────────────────────────────────────────────

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="4F4F58")
SUBHEADER_FILL = PatternFill("solid", fgColor="2A2A32")
THIN = Side(border_style="thin", color="888888")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


def _build_roster_sheet(wb: Workbook, db: Session, location: str, year: int) -> str:
    activity_ids = _activity_ids_for_location(db, location)
    home_churches = LOCATION_HOME_CHURCHES[location]

    # Members — active in the relevant home churches
    members = (
        db.query(
            Member.id,
            Member.full_name,
            Member.email,
            Member.category,
        )
        .filter(
            Member.member_status == "active",
            Member.home_church.in_(home_churches),
        )
        .order_by(Member.full_name)
        .all()
    )

    # Active iCare per member (for the "iCare" column)
    icare_rows = (
        db.query(IcareMember.member_id, IcareGroup.name)
        .join(IcareGroup, IcareGroup.id == IcareMember.icare_id)
        .filter(IcareMember.is_active.is_(True), IcareGroup.is_active.is_(True))
        .all()
    )
    icare_by_member = {r.member_id: r.name for r in icare_rows}

    # Attendance for the year — only for the relevant activities
    att_rows = []
    if activity_ids:
        att_rows = (
            db.query(
                Attendance.member_id,
                func.date(Attendance.timestamp).label("service_date"),
                Attendance.timestamp,
            )
            .filter(
                Attendance.activity_type_id.in_(activity_ids),
                func.extract("year", Attendance.timestamp) == year,
            )
            .all()
        )

    service_dates: set[date] = set()
    att_by_key: dict[tuple[int, date], datetime] = {}
    for r in att_rows:
        sd = r.service_date if isinstance(r.service_date, date) else r.timestamp.date()
        service_dates.add(sd)
        att_by_key[(r.member_id, sd)] = r.timestamp

    # Newest first to match Excel layout
    sorted_dates = sorted(service_dates, reverse=True)

    sheet_name = _unique_sheet_name(wb, f"Roster {year} {LOCATION_SHORT[location]}")
    ws = wb.create_sheet(sheet_name)

    # Title
    ws["A1"] = f"Sunday Service Roster — {LOCATION_SHORT[location]} {year}"
    ws["A1"].font = Font(bold=True, size=13)

    base_cols = ["No.", "Full name", "Email", "iCare", "Kategori"]

    # Date header row (spans Onsite + Online)
    header_row1 = 3
    header_row2 = 4

    for i, label in enumerate(base_cols, start=1):
        c = ws.cell(row=header_row1, column=i, value=label)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER
        ws.cell(row=header_row2, column=i, value="").border = BORDER
        ws.merge_cells(start_row=header_row1, start_column=i, end_row=header_row2, end_column=i)

    col_cursor = len(base_cols) + 1
    for sd in sorted_dates:
        # Date label spans two sub-columns: Onsite | Online
        c = ws.cell(row=header_row1, column=col_cursor, value=f"{sd.month}/{sd.day}")
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER
        ws.cell(row=header_row1, column=col_cursor + 1).border = BORDER
        ws.merge_cells(
            start_row=header_row1, start_column=col_cursor,
            end_row=header_row1, end_column=col_cursor + 1,
        )

        cs1 = ws.cell(row=header_row2, column=col_cursor,     value="Onsite")
        cs2 = ws.cell(row=header_row2, column=col_cursor + 1, value="Online")
        for cs in (cs1, cs2):
            cs.font = Font(bold=True, color="FFFFFF")
            cs.fill = SUBHEADER_FILL
            cs.alignment = CENTER
            cs.border = BORDER
        col_cursor += 2

    # Body rows
    body_start = header_row2 + 1
    for idx, m in enumerate(members, start=1):
        row = body_start + idx - 1
        ws.cell(row=row, column=1, value=idx).alignment = CENTER
        ws.cell(row=row, column=2, value=m.full_name)
        ws.cell(row=row, column=3, value=m.email)
        ws.cell(row=row, column=4, value=icare_by_member.get(m.id, "Belum mengikut"))
        ws.cell(row=row, column=5, value=_kategori(m.category))

        c = len(base_cols) + 1
        for sd in sorted_dates:
            ts = att_by_key.get((m.id, sd))
            onsite_val = ts.strftime("%H:%M") if ts else 0
            ws.cell(row=row, column=c,     value=onsite_val).alignment = CENTER
            ws.cell(row=row, column=c + 1, value=0).alignment = CENTER  # online placeholder
            c += 2

    # Column widths
    widths = [5, 28, 28, 22, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for c in range(len(base_cols) + 1, col_cursor):
        ws.column_dimensions[get_column_letter(c)].width = 7

    ws.freeze_panes = ws.cell(row=body_start, column=len(base_cols) + 1)
    return sheet_name


def _build_summary_sheet(wb: Workbook, db: Session, location: str, year: int) -> str:
    activity_ids = _activity_ids_for_location(db, location)

    # Weekly headcount per (service_date, kategori)
    weekly_rows = []
    weekly_total_rows = []
    if activity_ids:
        weekly_rows = (
            db.query(
                func.date(Attendance.timestamp).label("service_date"),
                Member.category.label("category"),
                func.count(Attendance.id).label("headcount"),
            )
            .join(Member, Member.id == Attendance.member_id)
            .filter(
                Attendance.activity_type_id.in_(activity_ids),
                func.extract("year", Attendance.timestamp) == year,
            )
            .group_by(func.date(Attendance.timestamp), Member.category)
            .all()
        )

        weekly_total_rows = (
            db.query(
                func.date(Attendance.timestamp).label("service_date"),
                func.count(Attendance.id).label("headcount"),
            )
            .filter(
                Attendance.activity_type_id.in_(activity_ids),
                func.extract("year", Attendance.timestamp) == year,
            )
            .group_by(func.date(Attendance.timestamp))
            .all()
        )

    # Aggregate: per-month and per-quarter buckets per category
    by_month_kat: dict[tuple[int, str], list[int]] = defaultdict(list)
    by_quarter_kat: dict[tuple[int, str], list[int]] = defaultdict(list)
    for r in weekly_rows:
        sd = r.service_date if isinstance(r.service_date, date) else r.service_date
        kat = _kategori(r.category)
        m = sd.month
        q = (sd.month - 1) // 3 + 1
        by_month_kat[(m, kat)].append(r.headcount)
        by_quarter_kat[(q, kat)].append(r.headcount)

    by_month_total: dict[int, list[int]] = defaultdict(list)
    by_quarter_total: dict[int, list[int]] = defaultdict(list)
    for r in weekly_total_rows:
        sd = r.service_date
        by_month_total[sd.month].append(r.headcount)
        by_quarter_total[(sd.month - 1) // 3 + 1].append(r.headcount)

    def avg(values: list[int]) -> int | str:
        return round(sum(values) / len(values)) if values else 0

    sheet_name = _unique_sheet_name(wb, f"Summary {year} {LOCATION_SHORT[location]}")
    ws = wb.create_sheet(sheet_name)

    short = LOCATION_SHORT[location]

    # ── Monthly section ──
    ws["A1"] = f"Monthly Average of: {short}"
    ws["A1"].font = Font(bold=True, size=13)

    monthly_header_row = 3
    headers = ["Month"] + [str(i) for i in range(1, 13)] + [f"Goal {year}"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=monthly_header_row, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER

    for ri, kat in enumerate(CATEGORY_ORDER, start=1):
        row = monthly_header_row + ri
        ws.cell(row=row, column=1, value=kat).font = Font(bold=True)
        for m in range(1, 13):
            ws.cell(row=row, column=1 + m, value=avg(by_month_kat.get((m, kat), []))).alignment = CENTER

    sub_row = monthly_header_row + len(CATEGORY_ORDER) + 1
    ws.cell(row=sub_row, column=1, value="Sub-Total").font = Font(bold=True)
    for m in range(1, 13):
        ws.cell(row=sub_row, column=1 + m, value=avg(by_month_total.get(m, []))).alignment = CENTER

    # ── Quarterly section ──
    q_title_row = sub_row + 3
    ws.cell(row=q_title_row, column=1, value=f"Quarterly Average of: {short}").font = Font(bold=True, size=13)

    q_header_row = q_title_row + 2
    q_headers = ["Quarterly", "1", "2", "3", "4", f"Goal {year}"]
    for i, h in enumerate(q_headers, start=1):
        c = ws.cell(row=q_header_row, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER

    for ri, kat in enumerate(CATEGORY_ORDER, start=1):
        row = q_header_row + ri
        ws.cell(row=row, column=1, value=kat).font = Font(bold=True)
        for q in range(1, 5):
            ws.cell(row=row, column=1 + q, value=avg(by_quarter_kat.get((q, kat), []))).alignment = CENTER

    q_sub_row = q_header_row + len(CATEGORY_ORDER) + 1
    ws.cell(row=q_sub_row, column=1, value="Sub-Total").font = Font(bold=True)
    for q in range(1, 5):
        ws.cell(row=q_sub_row, column=1 + q, value=avg(by_quarter_total.get(q, []))).alignment = CENTER

    # Column widths
    ws.column_dimensions["A"].width = 22
    for i in range(2, 15):
        ws.column_dimensions[get_column_letter(i)].width = 9

    return sheet_name


# ── Export endpoint ───────────────────────────────────────────────────────────

BUILDERS = {
    "sunday_attendance_roster":  _build_roster_sheet,
    "sunday_attendance_summary": _build_summary_sheet,
}


@router.post("/export")
def export_reports(
    payload: ReportExportRequest,
    db: Session = Depends(get_db),
    _=Depends(require_active_user),
):
    if not payload.reports:
        raise HTTPException(status_code=400, detail="No reports provided")

    wb = Workbook()
    # Drop the default empty sheet — we add one per report.
    wb.remove(wb.active)

    for item in payload.reports:
        builder = BUILDERS.get(item.type)
        if not builder:
            raise HTTPException(status_code=400, detail=f"Unknown report type: {item.type}")
        location = _validate_location(item.params.get("location", ""))
        year = _validate_year(item.params.get("year"))
        builder(wb, db, location, year)

    if not wb.sheetnames:
        # Defensive: openpyxl needs at least one sheet
        wb.create_sheet("Empty")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"reports_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
